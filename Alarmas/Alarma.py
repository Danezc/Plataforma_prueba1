import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image 
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


def creador_alarma(df_origen):
    """
    Esta función realiza las siguientes operaciones:
    
    1. Lee múltiples archivos Excel de un directorio especificado y los concatena en un único DataFrame (primer_df).
    2. Filtra el DataFrame primer_df según ciertos criterios en las columnas 'Causales' y 'AGO manual'.
    3. Crea un segundo DataFrame (segundo_df) con columnas seleccionadas y renombradas.
    4. Cambia el tipo de datos de ciertas columnas en segundo_df a string.
    5. Guarda segundo_df en un nuevo archivo Excel con ajustes de formato específicos.
    6. Ajusta el tamaño de las columnas al contenido.
    7. Personaliza el estilo de las celdas y agrega una imagen en el archivo Excel generado.

    Returns: None 

    """
    
    
    primer_df = df_origen

    maestro_df = pd.read_excel('Maestro Puntos de Venta.xlsx')

    primer_df = pd.merge(primer_df, maestro_df[['Gln Punto de Venta', 'Cadena']], 
                        on='Gln Punto de Venta', 
                        how='left')
    
    primer_df = pd.merge(primer_df, maestro_df[['Gln Punto de Venta', 'Ciudad']], 
                        on='Gln Punto de Venta', 
                        how='left')
    
    conteo_total_industrial = primer_df['Nombre Proveedor'].value_counts()

    conteo_total_cadena = primer_df['Cadena'].value_counts()

    # Filtrar las filas donde 'Causal' y 'AGO manual' son diferentes de 0 o NaN
    filtro = (primer_df['Causales'].notna() & primer_df['Causales'] != 0) & \
             (primer_df['AGO manual'].notna() & primer_df['AGO manual'] != 1)
    
    primer_df = primer_df[filtro]

    # Contar de nuevo los registros para cada "Nombre Proveedor" después de aplicar el filtro
    conteo_filtrado_industrial = primer_df['Nombre Proveedor'].value_counts()
    conteo_filtrado_cadena = primer_df['Cadena'].value_counts()

    # Calcular el indicador para cada "Nombre Proveedor"
    indicador_industrial = conteo_filtrado_industrial / conteo_total_industrial

    indicador_cadena = conteo_filtrado_cadena / conteo_total_cadena

    try:
    # Convertir la columna 'FECHA' a formato de fecha y hora
        primer_df['FECHA'] = pd.to_datetime(primer_df['FECHA']).dt.date
    except Exception as e:
        print("Error en formato de fecha:", e)
        
    fecha_primera_fila = primer_df['FECHA'].iloc[0].strftime("%Y-%m-%d")



    # Paso 2: Crear segundo_df con las columnas requeridas y llenarlo con base en primer_df
    column_map = {
        'FECHA': 'Fecha',
        'Nombre Proveedor': 'Industrial',
        'Ciudad': 'Ciudad',
        'Gln Punto de Venta': 'Gln Punto de Venta',
        'Punto Venta': 'Punto de Venta',
        'GTIN': 'Gtin',
        'Descripción PROD': 'Producto',
        'Causales': 'Causales',
    }
    
    # Mapea directamente las columnas que necesitas y crea segundo_df
    segundo_df = primer_df[list(column_map.keys())].rename(columns=column_map)
        # Cargar el DataFrame maestro desde 'Maestro Puntos de Venta.xlsx'

    # Realiza un merge con 'Gln Punto de Venta' como clave
    # para traer el valor correspondiente de la columna 'Formato'
    segundo_df = pd.merge(segundo_df, maestro_df[['Gln Punto de Venta', 'Cadena']], 
                        on='Gln Punto de Venta', 
                        how='left')

    # Renombra la columna 'Formato' a 'Cadena'
    segundo_df.rename(columns={'Cadena': 'Cadena'}, inplace=True)

    # Reorganiza las columnas
    column_order = ['Fecha', 'Industrial', 'Cadena', 'Ciudad','Gln Punto de Venta', 'Punto de Venta', 'Gtin', 'Producto','Causales']
    segundo_df = segundo_df[column_order]

    # Cambiar el tipo de datos a str para columnas específicas
    segundo_df['Gln Punto de Venta'] = segundo_df['Gln Punto de Venta'].astype(str)
    segundo_df['Cadena'] = segundo_df['Cadena'].astype(str)
    segundo_df['Gtin'] = segundo_df['Gtin'].astype(str)
    segundo_df['Fecha'] = pd.to_datetime(segundo_df['Fecha'], errors='coerce')
    segundo_df['Fecha'] = segundo_df['Fecha'].dt.strftime('%Y-%m-%d')

    # Crear la carpeta 
    carpeta_fecha = os.path.join('Alarma ' + fecha_primera_fila)
    if not os.path.exists(carpeta_fecha):
        os.makedirs(carpeta_fecha)
    Industriales_todos = []
    # Agrupamos por industrial y por Fecha y guarda
    for (gln, fecha), group in segundo_df.groupby(['Industrial', 'Fecha']):
        filename = f"Alarma_{group['Industrial'].iloc[0]}_{group['Fecha'].iloc[0]}.xlsx"
        # Reemplaza las barras con otro carácter
        filename = filename.replace("/", "-").replace("\\", "-")
        directorio_para_guardar = f'{carpeta_fecha}/Industriales'
        if not os.path.exists(directorio_para_guardar):
            os.makedirs(directorio_para_guardar)
        filepath = os.path.join(directorio_para_guardar, filename)
        # Guardar el DataFrame en un archivo Excel
        group.to_excel(filepath, index=False, startrow=6)

        # Añadir el indicador en la celda B6
        libro = openpyxl.load_workbook(filepath)
        hoja = libro.active
        
            # Obtener el indicador para el "industrial" actual
        industrial_actual = group['Industrial'].iloc[0]
        indicador_actual_industrial = indicador_industrial.get(industrial_actual, 0)  # Usar 0 como valor predeterminado si el industrial no se encuentra
        Industriales_todos.append(industrial_actual)
        # Insertar el indicador en la celda B6
        hoja['B6'] = indicador_actual_industrial
        hoja['A6'] = 'Su indicador es:'
        hoja['B6'].number_format = '0.00%'
        
        # Guardar los cambios
        libro.save(filepath)

    # Agrupa por cadena y por y guarda
    for (cadena, fecha), group in segundo_df.groupby(['Cadena', 'Fecha']):
        filename = f"Alarma_{group['Cadena'].iloc[0]}_{group['Fecha'].iloc[0]}.xlsx"
        # Reemplaza las barras con otro carácter
        filename = filename.replace("/", "-").replace("\\", "-")
        directorio_para_guardar = f'{carpeta_fecha}/Cadenas'
        if not os.path.exists(directorio_para_guardar):
            os.makedirs(directorio_para_guardar)
        filepath = os.path.join(directorio_para_guardar, filename)
        # Guardar el DataFrame en un archivo Excel
        group.to_excel(filepath, index=False, startrow=6)
        
        # Añadir el indicador en la celda B6
        libro = openpyxl.load_workbook(filepath)
        hoja = libro.active

        # Obtener el indicador para el "industrial" actual
        cadena_actual = group['Cadena'].iloc[0]
        indicador_actual_cadena = indicador_cadena.get(cadena_actual, 0)  # Usar 0 como valor predeterminado si el industrial no se encuentra
        
        # Insertar el indicador en la celda B6
        hoja['B6'] = indicador_actual_cadena
        hoja['A6'] = 'Su indicador es:'
        hoja['B6'].number_format = '0.00%'
        
        
        # Guardar los cambios
        libro.save(filepath)
        
    return carpeta_fecha, fecha_primera_fila


def ajustes_alarma(nombre_archivo):
    """
    Esta función aplica la función `ajustes_alarma` a todos los archivos Excel en un directorio
    específico para el día actual.

    Parámetros:
    - Ninguno.

    Notas:
    - Busca en los directorios 'Alarma {fecha_hoy}/Industriales' y 'Alarma {fecha_hoy}/Cadenas'
      donde fecha_hoy es la fecha actual.
    - Solo se aplica a archivos con extensión '.xlsx'.

    Retorno:
    - Ninguno, pero aplica los ajustes a todos los archivos Excel encontrados.
    """
    
    # Aeditar el archivo Excel para añadir la celda combinada y cambiar su color
    wb = openpyxl.load_workbook(nombre_archivo)
    ws = wb.active

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:  # En caso de que tenga error al encontrar una celda vacía
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Combinar celdas
    ws.merge_cells('C2:G4')


    # Establecer el borde naranja
    border = Border(
        left=Side(border_style='thin', color='ff4000'),
        right=Side(border_style='thin', color='ff4000'),
        top=Side(border_style='thin', color='ff4000'),
        bottom=Side(border_style='thin', color='ff4000')
    )

    # Establecer el color de fuente a naranja y usar la fuente Museo 900 en negrilla
    font = Font(name='Museo 900', bold=True, color="ff4500", size=20)

    # Centrar el texto en la celda
    alignment = Alignment(horizontal='center', vertical='center')

    # Iterar a través de cada celda en el rango C2:G4
    for row in ws['C2':'G4']:
        for cell in row:
            cell.border = border
            cell.font = font
            cell.alignment = alignment

    #Dialogo en la celda combinada 
    ws['C2'].value = 'PRODUCTOS AGOTADOS'
    ws['A6'].font = Font(color="ff4500", bold=True)
    ws['B6'].font = Font(bold=True)


    # Agregar imagen
    img = Image('D-LOGYCA_logo_excel.png')  
    img.width = img.width // 1  # Cambiar el ancho 
    img.height = img.height // 1  # Cambiar la altura

    # Posicionar la imagen en la celda A1
    ws.add_image(img, 'A1')

    # Guardar los cambios en el archivo Excel
    wb.save(nombre_archivo)
    print(f'Archivo guardado como {nombre_archivo}')

def aplicar_ajustes_a_todos_los_archivos(carpeta_fecha):
    """
    Aplica ajustes a todos los archivos de alarma en la carpeta de la fecha especificada.

    Args:
        carpeta_fecha (str): La ruta de la carpeta de la fecha.

    Returns:
        None
    """
    fecha_hoy = carpeta_fecha
    tipos_de_alarma = ['Industriales', 'Cadenas']  # Lista de directorios a considerar

    for tipo in tipos_de_alarma:
        directorio_alarma = f'{fecha_hoy}/{tipo}'

        if os.path.exists(directorio_alarma):
            for root, dirs, files in os.walk(directorio_alarma):
                for filename in files:
                    if filename.endswith('.xlsx'):
                        archivo_completo = os.path.join(root, filename)
                        ajustes_alarma(archivo_completo)
        else:
            print(f"Directorio {directorio_alarma} no encontrado.")

def Alarma_pdv(df_origen):
    """
    Procesa un archivo Excel de entrada, lo cruza con un maestro de puntos de venta
    para añadir la cadena correspondiente, mapea las columnas a un nuevo formato, y 
    genera archivos de Excel individuales agrupados por punto de venta.

    Parameters:
    entrada_path : tkinter.Entry
        Un objeto Entry de Tkinter que contiene la ruta del archivo Excel de entrada.

    Returns:
    str
        El path de la nueva carpeta creada donde se guardan los archivos generados.
    
    Raises:
    FileNotFoundError
        Si el archivo de entrada o el maestro de puntos de venta no se encuentra en la ruta especificada.
    Exception
        Si ocurre un error durante la lectura de los archivos Excel o durante el procesamiento de los datos.

    """
    primer_df = df_origen
    
    # Leer el archivo maestro de puntos de venta para obtener la cadena
    maestro_puntos_venta = pd.read_excel('Maestro Puntos de Venta.xlsx')
    
    # Cruzar la información para obtener la cadena
    merged_df = primer_df.merge(maestro_puntos_venta[['Gln Punto de Venta', 'Cadena']], on='Gln Punto de Venta', how='left')

    # Mapeo de columnas de entrada a salida
    merged_df['Fecha Medición'] = merged_df['FECHA']
    merged_df['Punto de Venta'] = merged_df['Punto Venta']
    merged_df['gtinproducto'] = merged_df['GTIN']
    # PLU se mantiene igual, asumiendo que ya está en el primer_df
    merged_df['descripcionproducto'] = merged_df['Descripción PROD']
    merged_df['Industrial'] = merged_df['Nombre Proveedor']
    merged_df['Novedad'] = merged_df['Causales']

    # Selección de las columnas para el archivo de salida
    merged_df = merged_df[['Fecha Medición', 'Punto de Venta', 'Cadena', 'gtinproducto', 'PLU', 'descripcionproducto', 'Industrial', 'Novedad']]

    # Agrupar por 'Punto de Venta'
    grouped = merged_df.groupby('Punto de Venta')

    # Crear la nueva carpeta
    fecha_primera_fila = primer_df['FECHA'].iloc[0].strftime("%Y-%m-%d")
    new_folder_path = f'Alarma PDV {fecha_primera_fila}'
    os.makedirs(new_folder_path, exist_ok=True)

    # Iterar sobre cada grupo y crear un archivo Excel
    for name, group in grouped:
        group = group.dropna(subset=['Novedad'])
        # Comprobar si el DataFrame agrupado está vacío
        if group.empty:
            print(f"No se ha creado el DF para el Punto de Venta {name}")
            continue  # Salta a la siguiente iteración del bucle

        fecha_capturador = group['Fecha Medición'].iloc[0].strftime('%Y-%m-%d')
        nombre_cadena = group['Cadena'].iloc[0]
        file_name = f'Alarma diaria_{fecha_capturador}_{name}_{nombre_cadena}.xlsx'
        file_path = os.path.join(new_folder_path, file_name)

        # Crear un nuevo libro de trabajo y seleccionar la hoja activa
        wb = Workbook()
        ws = wb.active
        ws.title = fecha_capturador

        # Definir las columnas que queremos en el nuevo archivo
        new_columns = ['Fecha Medición', 'Punto de Venta', 'Cadena', 'gtinproducto', 'PLU', 'descripcionproducto', 'Industrial', 'Novedad']

        # Escribir las cabeceras en la fila 8 comenzando por la columna B
        for col_num, column_title in enumerate(new_columns, start=2):
            ws.cell(row=8, column=col_num).value = column_title

        # Escribir los datos en el archivo Excel
        for row_index, row in enumerate(group.itertuples(index=False), start=9):  # Empezar en la fila 9
            for col_num, column_title in enumerate(new_columns, start=2):
                column_value = getattr(row, column_title.replace(' ', '_'), None)  # Reemplaza espacios por guiones bajos
                if column_value is None:  # Si getattr no encuentra el atributo, usa la columna directamente
                    column_value = row[group.columns.get_loc(column_title)]
                ws.cell(row=row_index, column=col_num).value = column_value

        # Guardar el libro de trabajo
        wb.save(file_path)
        print(f"Procesado: {file_name}")  # Imprimir el nombre del archivo procesado

    return new_folder_path, fecha_primera_fila  # Retorna el path de la nueva carpeta creada


def ajustar_formato_excel(wb, folder_path_alarmas, folder_path_principal):
    """
    Ajusta el formato de la hoja de cálculo en el libro de trabajo proporcionado.

    Parámetros:
    - wb: openpyxl.workbook.Workbook
        El libro de trabajo de Excel que se va a procesar.
    - folder_path: str
        La ruta de la carpeta donde se encuentra la subcarpeta 'Logos_Cadenas'.
    """

    folder_path_principal = folder_path_principal.get()
    estilo_fecha = NamedStyle(name='estilo_fecha', number_format='YYYY-MM-DD')
    script_folder_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Recorre todas las hojas del libro de trabajo
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Ajustar el ancho de las columnas
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        
               # Insertar imagen de encabezado
        path_to_logo = os.path.join(script_folder_path, 'Logyca_encabezado.png')
        if os.path.exists(path_to_logo):
            logo = Image(path_to_logo)
            # Configurar el tamaño de la imagen
            # Puedes ajustar estos valores según sea necesario
            logo.width = 1100
            logo.height = 100
            ws.add_image(logo, 'B2')

        # Insertar imagen correspondiente a la cadena
        cadena = ws.cell(row=9, column=4).value
        cadena_img_path = os.path.join(script_folder_path, 'Logos_Cadenas', f'{cadena}.png')

        if os.path.exists(cadena_img_path):
            cadena_img = Image(cadena_img_path)
            # Configurar el tamaño de la imagen
            cadena_img.width = 130  # Configura el ancho de la imagen
            cadena_img.height = 100  # Configura el alto de la imagen
            ws.add_image(cadena_img, 'I2')
        else:
            print(f'No se encontró la imagen para {cadena}')

        # Ajustar formato de gtinproducto como string
        for row in ws.iter_rows(min_row=9, max_col=ws.max_column, max_row=ws.max_row):
            gtin_celda = row[4]  
            plu_celda = row[5]
            if gtin_celda.value is not None:
                gtin_celda.value = str(gtin_celda.value)
            if plu_celda.value is not None:
                plu_celda.value = str(plu_celda.value)

            # Iterar sobre las celdas en la columna B comenzando desde la fila 9
        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=2, max_col=2):  # Columna B
            for cell in row:
                if cell.value:  # Verificar si la celda tiene un valor
                    cell.style = estilo_fecha
        # Añadir bordes y aplicar negrita a los encabezados
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                if cell.row == 8:
                    cell.font = Font(bold=True)
                    ws.auto_filter.ref = ws.cell(row=cell.row, column=cell.col_idx).coordinate

                        
    return wb  # Retorna el libro de trabajo ajustado

def procesar_archivos_formato(folder_path_alarmas, folder_path_principal):
    """
    Procesa todos los archivos Excel en la carpeta dada, ajustando el formato de cada uno.
    
    Parámetros:
    - folder_path : str
        La ruta de la carpeta que contiene los archivos Excel de salida para formatear.
    """
    for file_name in os.listdir(folder_path_alarmas):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path_alarmas, file_name)
            wb = load_workbook(file_path)
            ajustar_formato_excel(wb, folder_path_alarmas, folder_path_principal)
            wb.save(file_path)
            print(f'Archivo formateado: {file_name}')

